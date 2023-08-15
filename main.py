import os
import json
import zipfile
import requests
import openpyxl
import pandas as pd
import logging
from pathlib import Path
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import as_completed

# Configure logging
logging.basicConfig(level=logging.INFO)

# Function to load configuration
def load_config():
    with open('config.json') as file:
        return json.load(file)

# Function to download excel file
def download_excel(url, headers):
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        # Define the file path
        file_path = 'data.xlsx'  # This saves the file in the current working directory

        with open(file_path, 'wb') as file:
            file.write(response.content)

        logging.info("Excel file downloaded.")
    except requests.RequestException as e:
        logging.error(f"Error downloading Excel file: {e}")

def download_image(session, url, path, headers):
    try:
        response = session.get(url, headers=headers)
        response.raise_for_status()
        with open(path, 'wb') as file:
            file.write(response.content)
    except requests.RequestException as e:
        logging.error(f"Error downloading image {url}: {e}")
        return url  # Return the URL if there's an error

# Function to process excel file
def process_excel(config):
    workbook = openpyxl.load_workbook('data.xlsx')

    # List to store failed URLs
    failed_urls = []

    # Create a Session object to reuse the connection
    session = requests.Session()

    # Iterate through sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        Path(f'images/{sheet_name}').mkdir(parents=True, exist_ok=True) #creates a directory for each sheet
        logging.info(f"Processing sheet: {sheet_name}")

        # Use ThreadPoolExecutor for parallel downloads
        with ThreadPoolExecutor() as executor:
            tasks = []
            # Iterate through rows to find columns with _URL in the title
            for col in sheet.iter_cols():
                col_name = col[0].value
                if col_name and '_URL' in col_name:
                    for cell_idx, cell in enumerate(col[1:]):
                        if cell.value and (cell.value.startswith('http://') or cell.value.startswith('https://')):
                            image_url = cell.value
                            image_name = os.path.join('images', sheet_name, f"{sheet.cell(row=cell_idx+2, column=1).value}")

                            # Download and save the image
                            tasks.append(executor.submit(download_image, session, image_url, image_name, config['headers']))

            # Track progress using tqdm and log failed URLs
            for future in tqdm(as_completed(tasks), desc=f"Downloading images from {sheet_name}", total=len(tasks), unit="file"):
                result = future.result()
                if result:  # If a URL was returned (download failed), log it
                    log_failed_url(sheet_name, cell, result, failed_urls)

        # Zip the folder
        Path('zip').mkdir(exist_ok=True)
        zip_path = os.path.join('zip', f'{sheet_name}.zip')
        with zipfile.ZipFile(zip_path, 'w') as zip_ref:
            for root, _, files in os.walk(os.path.join('images', sheet_name)):
                for file in files:
                    zip_ref.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), os.path.join('images', sheet_name)))

        logging.info(f"Zipped folder created: {sheet_name}.zip")

    # Save the failed URLs to an Excel file
    save_failed_urls_to_excel(failed_urls)
        
# Function to log failed URLs
def log_failed_url(sheet_name, cell, url, failed_urls):
    failed_urls.append({'Sheet': sheet_name, 'Cell': cell.coordinate, 'URL': url})

# Function to save failed URLs to an Excel file
def save_failed_urls_to_excel(failed_urls):
    failed_urls_df = pd.DataFrame(failed_urls, columns=['Sheet', 'Cell', 'URL'])
    failed_urls_df.to_excel('failed_urls.xlsx', index=False)
    logging.info(f"Failed URLs saved to 'failed_urls.xlsx'")

# Main function
def main():
    config = load_config()
    url = config["project"]
    config['headers'] = {'Authorization': f'Token {config["api_key"]}'}
    download_excel(url, config['headers'])
    process_excel(config)

if __name__ == "__main__":
    main()
